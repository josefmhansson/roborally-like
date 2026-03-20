from __future__ import annotations

import argparse
import json
import subprocess
import sys
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / 'specifications.xlsx'

CARD_TYPE_LABELS = {
    'reinforcement': 'reinforce',
    'movement': 'mov',
    'attack': 'atk',
    'spell': 'spell',
}

CLASS_LABELS = {
    'commander': 'Commander',
    'warleader': 'Warleader',
    'archmage': 'Archmage',
    None: 'All',
}

ENCOUNTERS = [
    {
        'name': 'Slimes',
        'units': '1x Grandslime (5 + floor(n/2)); floor(n/4)x Slime (3 + floor(n/4)); floor(n/3)x Slimeling (1 + floor(n/8))',
        'ap_budget': '3 + floor(n/5)',
        'deck': [
            'move_forward_face',
            'move_forward_face',
            'move_forward_face',
            'move_forward_face',
            'attack_coordinated',
            'attack_coordinated',
            'attack_roguelike_basic',
            'attack_roguelike_basic',
            'attack_roguelike_basic',
            'attack_roguelike_basic',
            'move_tandem',
            'move_tandem',
        ],
    },
    {
        'name': 'Trolls',
        'units': '2 + floor(n/5)x Troll (10 + floor(n/2))',
        'ap_budget': '3 + floor(n/5)',
        'deck': [
            'move_forward_face',
            'move_forward_face',
            'move_forward_face',
            'attack_roguelike_slow',
            'attack_roguelike_slow',
            'attack_roguelike_slow',
            'attack_roguelike_slow',
            'attack_roguelike_stomp',
            'attack_roguelike_stomp',
            'attack_roguelike_stomp',
        ],
    },
    {
        'name': 'Wolf Pack',
        'units': '1x Alpha Wolf (4 + floor(n/3)); 4x Wolf (2 + floor(n/6))',
        'ap_budget': '3 + floor(n/5)',
        'deck': [
            'move_forward',
            'move_forward',
            'move_forward',
            'move_forward',
            'attack_roguelike_basic',
            'attack_roguelike_basic',
            'attack_roguelike_basic',
            'attack_roguelike_basic',
            'attack_roguelike_pack_hunt',
            'attack_roguelike_pack_hunt',
            'spell_roguelike_mark',
            'spell_roguelike_mark',
        ],
    },
    {
        'name': 'Ice Spirits',
        'units': '3x Ice Spirit (2 + floor(n/3))',
        'ap_budget': '3 + floor(n/5)',
        'deck': [
            'attack_ice_bolt',
            'attack_ice_bolt',
            'attack_ice_bolt',
            'attack_ice_bolt',
            'spell_blizzard',
            'move_forward',
            'move_forward',
            'move_forward',
            'move_forward',
            'reinforce_roguelike_split',
        ],
    },
    {
        'name': 'Fire Spirits',
        'units': '3x Fire Spirit (2 + floor(n/3))',
        'ap_budget': '3 + floor(n/5)',
        'deck': [
            'attack_fireball',
            'attack_fireball',
            'attack_line',
            'attack_line',
            'spell_meteor',
            'move_forward',
            'move_forward',
            'move_any',
            'move_any',
            'reinforce_roguelike_split',
        ],
    },
    {
        'name': 'Lightning Spirits',
        'units': '3x Lightning Spirit (2 + floor(n/3))',
        'ap_budget': '3 + floor(n/5)',
        'deck': [
            'attack_chain_lightning',
            'attack_chain_lightning',
            'attack_chain_lightning',
            'attack_chain_lightning',
            'spell_roguelike_thunderstorm',
            'move_forward',
            'move_forward',
            'move_forward',
            'move_forward',
            'reinforce_roguelike_split',
        ],
    },
    {
        'name': 'Bandits',
        'units': '5x Bandit (3 + floor(n/5))',
        'ap_budget': '3 + floor(n/5)',
        'deck': [
            'move_forward',
            'move_forward',
            'move_forward',
            'move_any',
            'move_any',
            'move_any',
            'move_pivot',
            'move_pivot',
            'move_double_steps',
            'move_double_steps',
            'attack_fwd_lr',
            'attack_fwd_lr',
            'attack_arrow',
            'attack_arrow',
            'attack_charge',
            'attack_whirlwind',
            'attack_blade_dance',
            'attack_execute',
            'reinforce_boost',
            'reinforce_boost',
            'spell_snare',
            'attack_jab',
            'attack_jab',
            'attack_shove',
        ],
    },
    {
        'name': 'Necromancer',
        'units': '1x Necromancer (4 + floor(n/4)); 2 + floor(n/4)x Skeleton Soldier/Warrior/Mage (2)',
        'ap_budget': '3 + floor(n/5)',
        'deck': [
            'spell_roguelike_raise',
            'spell_roguelike_raise',
            'attack_ice_bolt',
            'attack_ice_bolt',
            'attack_coordinated',
            'attack_roguelike_basic',
            'attack_roguelike_basic',
            'move_tandem',
            'move_any',
            'move_any',
            'move_any',
            'move_forward',
            'move_forward',
            'spell_brain_freeze',
        ],
    },
]

MONSTER_UNITS = [
    {
        'name': 'Grandslime',
        'group': 'Slimes',
        'health': '5 + floor(n/2)',
        'assets': 'assets/monsters/monster_grandslime.png',
        'traits': 'Splits into 2 Slimes on death.',
    },
    {
        'name': 'Slime',
        'group': 'Slimes',
        'health': '3 + floor(n/4)',
        'assets': 'assets/monsters/monster_slime.png',
        'traits': 'Splits into 2 Slimelings on death.',
    },
    {
        'name': 'Slimeling',
        'group': 'Slimes',
        'health': '1 + floor(n/8)',
        'assets': 'assets/monsters/monster_slimeling.png',
        'traits': 'Small slime spawned by larger slimes.',
    },
    {
        'name': 'Troll',
        'group': 'Trolls',
        'health': '10 + floor(n/2)',
        'assets': 'assets/monsters/monster_troll.png',
        'traits': 'Large brute used by the Trolls encounter.',
    },
    {
        'name': 'Alpha Wolf',
        'group': 'Wolf Pack',
        'health': '4 + floor(n/3)',
        'assets': 'assets/monsters/monster_alpha_wolf.png',
        'traits': 'Only unit that can use Pack Hunt.',
    },
    {
        'name': 'Wolf',
        'group': 'Wolf Pack',
        'health': '2 + floor(n/6)',
        'assets': 'assets/monsters/monster_wolf_2.png',
        'traits': 'Pack unit that supports Alpha Wolf.',
    },
    {
        'name': 'Ice Spirit',
        'group': 'Ice Spirits',
        'health': '2 + floor(n/3)',
        'assets': 'assets/monsters/monster_ice_elemental.png',
        'traits': 'Encounter pressures player units with Chilled and Slow synergies.',
    },
    {
        'name': 'Fire Spirit',
        'group': 'Fire Spirits',
        'health': '2 + floor(n/3)',
        'assets': 'assets/monsters/monster_fire_elemental.png',
        'traits': 'Has Scalding: attack damage also applies Burn.',
    },
    {
        'name': 'Lightning Spirit',
        'group': 'Lightning Spirits',
        'health': '2 + floor(n/3)',
        'assets': 'assets/monsters/monster_lightning_elemental.png',
        'traits': 'Has Lightning Barrier.',
    },
    {
        'name': 'Bandit',
        'group': 'Bandits',
        'health': '3 + floor(n/5)',
        'assets': '\n'.join(
            [
                'assets/monsters/monster_bandit_1.png',
                'assets/monsters/monster_bandit_2.png',
                'assets/monsters/monster_bandit_3.png',
            ]
        ),
        'traits': 'Humanoid encounter unit with three art variants.',
    },
    {
        'name': 'Necromancer',
        'group': 'Necromancer',
        'health': '4 + floor(n/4)',
        'assets': 'assets/monsters/monster_necromancer.png',
        'traits': 'Slow. Spell Resistance. Reinforcement gains halved. Raises skeletons from fallen player units.',
    },
    {
        'name': 'Skeleton Soldier',
        'group': 'Necromancer',
        'health': 2,
        'assets': 'assets/monsters/monster_skeleton_soldier.png',
        'traits': 'Minion skeleton raised by Necromancer.',
    },
    {
        'name': 'Skeleton Warrior',
        'group': 'Necromancer',
        'health': 2,
        'assets': 'assets/monsters/monster_skeleton_warrior.png',
        'traits': 'Minion skeleton raised by Necromancer.',
    },
    {
        'name': 'Skeleton Mage',
        'group': 'Necromancer',
        'health': 2,
        'assets': 'assets/monsters/monster_skeleton_mage.png',
        'traits': 'Minion skeleton raised by Necromancer.',
    },
]

STATUS_ROWS = [
    ('Burn', 'Debuff', 'Deals 1 damage to the unit at end of turn.', 'Multiple applications currently collapse back to one remaining Burn entry after ticking.'),
    ('Cannot Move', 'Debuff', 'Unit cannot move while the effect lasts.', 'Used by Trip, Snare, and bear traps.'),
    ('Stunned', 'Debuff', 'Unit cannot act while the effect lasts.', 'Applied by Bash, Petrify, and Stomp.'),
    ('Slow', 'Debuff', 'Unit movement is limited to 1 tile total per turn.', 'Also used as a card keyword on some roguelike attacks.'),
    ('Chilled', 'Debuff', 'Counts as a movement-slowing debuff.', 'If a unit is both Chilled and Slowed it also becomes Frozen.'),
    ('Frozen', 'Debuff', 'Unit cannot move or act.', 'Derived from Chilled + Slow.'),
    ('Spell Resistance', 'Buff', 'Spell damage to this unit is halved, rounded down.', 'Base leader modifier and also used by Necromancer.'),
    ('Reinforcement Penalty', 'Debuff', 'Strength gains from reinforcement effects are halved, rounded down.', 'Base leader modifier and also used by Necromancer.'),
    ('Regeneration', 'Buff', 'Heals 1 strength at end of turn per stack.', 'Present in engine support and tests.'),
    ('Disarmed', 'Debuff', 'Reduces damage dealt by 1 per stack.', 'Applied by Disarm.'),
    ('Vulnerable', 'Debuff', 'Increases damage taken by 1 per stack.', 'Applied by Bleed and Rage; stacks.'),
    ('Strong', 'Buff', 'Increases damage dealt by 1 per stack.', 'Stacks; Commander aura also grants an indefinite Strong source.'),
    ('Undying', 'Buff', 'Prevents damage and ignores incoming debuffs while active.', 'Used by Shrug Off and Berserk.'),
    ('Spikes', 'Buff', 'Reflects 1 damage to the attacker when this unit takes damage.', 'Used by Spikes.'),
    ('Marked for Death', 'Mixed', 'Unit is destroyed at the end of the turn.', 'Renamed from Berserk. Paired with Strong and Undying for the same turn on the Berserk card.'),
    ('Scalding', 'Buff', 'Attack damage from this unit also applies Burn.', 'Encounter-only modifier used by Fire Spirits.'),
    ('Lightning Barrier', 'Buff', 'Deals 1 damage to each adjacent enemy at end of turn.', 'Encounter-only on Lightning Spirits and granted by the Lightning Barrier card.'),
]

MISC_ROWS = [
    (None, 'Classes'),
    (None, 'Commander: Soldier unit. Leader grants adjacent friendly units Strong.'),
    (None, 'Warleader: Warrior unit. Leader is not Slowed by default.'),
    (None, 'Archmage: Mage unit. Leader grants +1 AP next turn if it held position.'),
    (None, 'Document'),
    (None, 'n in encounter/unit formulas means roguelike match number, starting at 1.'),
    (None, 'Spreadsheet rows are intended to stay aligned with player-facing content in code.'),
]


def extract_code_data() -> dict:
    script = """
import { CARD_DEFS } from './src/engine/cards.ts'
import { PLAYER_CLASS_DEFS } from './src/engine/classes.ts'

const cards = Object.values(CARD_DEFS).map((card) => ({
  id: card.id,
  name: card.name,
  actionCost: card.actionCost ?? 1,
  type: card.type,
  classId: card.classId ?? null,
  description: card.description,
  keywords: card.keywords ?? [],
  roguelikeOnly: card.roguelikeOnly === true,
  countsAs: card.countsAs ?? [],
  effectTypes: card.effects.map((effect) => effect.type),
  canTargetBarricades: card.canTargetBarricades ?? null,
}))

console.log(JSON.stringify({
  cards,
  classes: Object.values(PLAYER_CLASS_DEFS),
}))
"""
    completed = subprocess.run(
        'npx tsx -',
        cwd=ROOT,
        input=script,
        capture_output=True,
        text=True,
        check=True,
        shell=True,
    )
    return json.loads(completed.stdout)


def collapse_card_names(card_ids: list[str], card_names: dict[str, str]) -> str:
    counts = Counter(card_ids)
    ordered_ids: list[str] = []
    for card_id in card_ids:
        if card_id not in ordered_ids:
            ordered_ids.append(card_id)
    return ', '.join(f"{counts[card_id]}x {card_names[card_id]}" for card_id in ordered_ids)


def animation_for_card(card: dict) -> str:
    special = {
        'attack_arrow': 'Arrow projectile',
        'attack_ice_bolt': 'Ice bolt projectile',
        'attack_fireball': 'Fireball projectile and splash',
        'attack_line': 'Flame line',
        'attack_chain_lightning': 'Chain lightning arcs',
        'spell_roguelike_thunderstorm': 'Chain lightning arcs',
        'spell_lightning': 'Lightning strike',
        'spell_meteor': 'Meteor impact and blast',
        'spell_blizzard': 'Frost burst over radius',
        'spell_pitfall_trap': 'Trap placement marker',
        'spell_explosive_trap': 'Trap placement marker',
        'move_teleport': 'Teleport flash',
        'attack_volley': 'Volley projectiles',
        'attack_harpoon': 'Harpoon pull',
        'attack_whirlwind': 'Whirlwind hit and push',
    }
    if card['id'] in special:
        return special[card['id']]

    effect_types = set(card['effectTypes'])
    if 'teleport' in effect_types:
        return 'Teleport flash'
    if 'placeTrap' in effect_types:
        return 'Trap placement marker'
    if 'damageTileArea' in effect_types:
        return 'Meteor impact and blast'
    if 'damageRadius' in effect_types or 'lineSplash' in effect_types:
        return 'Area spell burst'
    if 'chainLightning' in effect_types or 'chainLightningAllFriendly' in effect_types:
        return 'Chain lightning arcs'
    if 'spawn' in effect_types or 'spawnAdjacentFriendly' in effect_types or 'spawnSkeletonAdjacent' in effect_types:
        return 'Spawn effect'
    if 'splitUnit' in effect_types:
        return 'Split and spawn effect'
    if 'boost' in effect_types or 'boostAllFriendly' in effect_types:
        return 'Buff pulse'
    if 'applyUnitModifier' in effect_types or 'clearUnitDebuffs' in effect_types or 'clearUnitModifiers' in effect_types:
        return 'Status effect pulse'
    if 'move' in effect_types or 'moveToTile' in effect_types or 'convergeTowardTile' in effect_types:
        return 'Moving unit'
    if card['type'] == 'attack':
        return 'Attack impact'
    if card['type'] == 'spell':
        return 'Spell effect'
    return 'Card effect'


def notes_for_card(card: dict) -> str:
    parts = [f"Id: {card['id']}"]
    if card['keywords']:
        parts.append('Keywords: ' + ', '.join(card['keywords']))
    if card['countsAs']:
        parts.append('Counts as: ' + ', '.join(card['countsAs']))
    if card['roguelikeOnly']:
        parts.append('Roguelike only')
    if card['canTargetBarricades'] is False:
        parts.append('Cannot target barricades')
    return '; '.join(parts)


def unit_rows(classes: list[dict]) -> list[list[object]]:
    rows: list[list[object]] = [['Name', 'Encounter/Class', 'Default Health', 'Asset(s)', 'Traits']]
    unit_traits = {
        'Commander': 'Base class unit. Gets Strong while adjacent to the Commander leader.',
        'Warleader': 'Base class unit.',
        'Archmage': 'Base class unit.',
    }
    leader_traits = {
        'Commander': 'Leader. Slow by default. Spell Resistance. Reinforcement gains halved. Adjacent friendly units gain Strong.',
        'Warleader': 'Leader. No default Slow. Spell Resistance. Reinforcement gains halved.',
        'Archmage': 'Leader. Slow by default. Spell Resistance. Reinforcement gains halved. Grants +1 AP next turn if it held position.',
    }

    for class_def in classes:
        class_name = class_def['name']
        rows.append(
            [
                class_def['unitName'],
                class_name,
                2,
                '\n'.join([class_def['unitBaseAsset'], class_def['unitTeamAsset']]),
                unit_traits[class_name],
            ]
        )
        rows.append(
            [
                class_name,
                class_name,
                5,
                '\n'.join([class_def['leaderBaseAsset'], class_def['leaderTeamAsset']]),
                leader_traits[class_name],
            ]
        )

    rows.append(
        [
            'Barricade',
            'Commander',
            1,
            '\n'.join(['assets/units/unit_barricade_base.png', 'assets/units/unit_barricade_team.png']),
            'Summoned structure created by the Barricade card.',
        ]
    )

    for unit in MONSTER_UNITS:
        rows.append([unit['name'], unit['group'], unit['health'], unit['assets'], unit['traits']])

    return rows


def build_tables(data: dict) -> dict[str, list[list[object]]]:
    cards = data['cards']
    classes = data['classes']
    card_names = {card['id']: card['name'] for card in cards}

    card_rows: list[list[object]] = [['Name', 'AP', 'Type', 'Class', 'Description', 'Animation', 'Notes']]
    for card in cards:
        card_rows.append(
            [
                card['name'],
                card['actionCost'],
                CARD_TYPE_LABELS[card['type']],
                CLASS_LABELS[card['classId']],
                card['description'],
                animation_for_card(card),
                notes_for_card(card),
            ]
        )

    encounter_rows: list[list[object]] = [['Name', 'Units', 'AP budget', 'Deck']]
    for encounter in ENCOUNTERS:
        encounter_rows.append(
            [
                encounter['name'],
                encounter['units'],
                encounter['ap_budget'],
                collapse_card_names(encounter['deck'], card_names),
            ]
        )

    return {
        'Cards': card_rows,
        'Encounters': encounter_rows,
        'Units': unit_rows(classes),
        'Misc': [[None, None], *MISC_ROWS],
        'Status effects': [['Name', 'Buff/debuff', 'Description', 'Notes'], *STATUS_ROWS],
    }


def reset_sheet(worksheet, rows: list[list[object]]) -> None:
    if worksheet.max_row > 0:
        worksheet.delete_rows(1, worksheet.max_row)
    for row in rows:
        worksheet.append(row)


def normalize_sheet_rows(worksheet, expected_width: int) -> list[list[object]]:
    rows: list[list[object]] = []
    for row in worksheet.iter_rows(values_only=True):
        values = list(row[:expected_width])
        while values and values[-1] is None:
            values.pop()
        rows.append(values + [None] * (expected_width - len(values)))

    while rows and all(value is None for value in rows[-1]):
        rows.pop()
    return rows


def compare_tables(workbook_path: Path, expected_tables: dict[str, list[list[object]]]) -> list[str]:
    workbook = openpyxl.load_workbook(workbook_path)
    errors: list[str] = []

    for sheet_name, expected_rows in expected_tables.items():
        if sheet_name not in workbook.sheetnames:
            errors.append(f"Missing sheet: {sheet_name}")
            continue
        worksheet = workbook[sheet_name]
        width = max(len(row) for row in expected_rows)
        actual_rows = normalize_sheet_rows(worksheet, width)
        normalized_expected = [list(row) + [None] * (width - len(row)) for row in expected_rows]

        if actual_rows == normalized_expected:
            continue

        errors.append(f"Sheet mismatch: {sheet_name}")
        max_rows = max(len(actual_rows), len(normalized_expected))
        for index in range(max_rows):
            actual = actual_rows[index] if index < len(actual_rows) else None
            expected = normalized_expected[index] if index < len(normalized_expected) else None
            if actual == expected:
                continue
            errors.append(f"  row {index + 1}: expected {expected}, found {actual}")
            if len(errors) >= 12:
                errors.append('  ...')
                return errors
    return errors


def write_tables(workbook_path: Path, expected_tables: dict[str, list[list[object]]]) -> None:
    workbook = openpyxl.load_workbook(workbook_path)
    for sheet_name, rows in expected_tables.items():
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
        reset_sheet(worksheet, rows)
    workbook.save(workbook_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Sync or audit specifications.xlsx against current game definitions.')
    parser.add_argument('mode', choices=['write', 'check'], nargs='?', default='check')
    parser.add_argument('--workbook', default=str(WORKBOOK_PATH))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        print(f'Workbook not found: {workbook_path}', file=sys.stderr)
        return 1

    data = extract_code_data()
    tables = build_tables(data)

    if args.mode == 'write':
        write_tables(workbook_path, tables)
        print(f'Updated {workbook_path}')
        return 0

    errors = compare_tables(workbook_path, tables)
    if errors:
        print('\n'.join(errors), file=sys.stderr)
        return 1

    print(f'{workbook_path} is in sync with the generated content tables.')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
